# =====================================
# Streamlit App: 提成表多sheet自动审核（总 + 轻卡 + 重卡）
# 标红错误格 + 标黄合同号 + 精简错误下载 + 独立错误数统计
# =====================================
import streamlit as st
import pandas as pd
from io import BytesIO
import unicodedata, re
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

st.title("📊 模拟人事用薪资计算表自动审核系统-3")

st.image("image/app3.png")

# ========== 上传文件 ==========
uploaded_files = st.file_uploader(
    "请上传文件名中包含以下字段的文件：“提成”、“放款明细”、“二次明细”和“原表”的xlsx文件。最后誊写，需检的表为提成表。使用2数据时，本司-1月格式的表需按照要求在文件名中加入”提成”，“原表”等区分。",
    type="xlsx", accept_multiple_files=True
)

if not uploaded_files or len(uploaded_files) < 4:
    st.warning("⚠️ 请至少上传 提成表、放款明细、二次明细、原表 四个文件")
    st.stop()

# ========== 工具函数 ==========
def find_file(files_list, keyword):
    for f in files_list:
        if keyword in f.name:
            return f
    return None

def normalize_text(val):
    if pd.isna(val):
        return ""
    s = str(val)
    s = re.sub(r'[\n\r\t ]+', '', s)
    s = s.replace('\u3000', '')
    return ''.join(unicodedata.normalize('NFKC', ch) for ch in s).lower().strip()

def normalize_num(val):
    if pd.isna(val):
        return None
    s = str(val).replace(",", "").strip() # <--- 1. 不再替换 "%"
    if s in ["", "-", "nan"]:
        return None
    try:
        # 2. 在这里检查和处理 "%"
        if "%" in s:
            s = s.replace("%", "")
            return float(s) / 100
        return float(s)
    except:
        return s

def find_col(df_like, keyword, exact=False):
    key = keyword.strip().lower()
    columns = df_like.columns if hasattr(df_like, "columns") else df_like.index
    for col in columns:
        cname = str(col).strip().lower()
        if (exact and cname == key) or (not exact and key in cname):
            return col
    return None

def normalize_contract_key(series: pd.Series) -> pd.Series:
    """
    对合同号 Series 进行标准化处理，用于安全的 pd.merge 操作。
    """
    s = series.astype(str)
    s = s.str.replace(r"\.0$", "", regex=True) 
    s = s.str.strip()
    s = s.str.upper() 
    s = s.str.replace('－', '-', regex=False)
    # (这个版本不移除内部空格，因为合同号可能包含它们)
    return s

def prepare_ref_df(df_list, required_cols_dict, prefix):
    """
    (新 V2) 预处理参考DF列表：合并、标准化Key、提取列、重命名
    required_cols_dict: {'合同': ('合同', False), '类型': ('类型', True), ...}
    """
    if not df_list or all(df is None for df in df_list):
        st.warning(f"⚠️ {prefix} 数据列表为空，跳过预处理。")
        return pd.DataFrame(columns=['__KEY__'])
        
    try:
        df_concat = pd.concat([df for df in df_list if df is not None], ignore_index=True)
    except Exception as e:
        st.error(f"❌ 预处理 {prefix} 时合并失败: {e}")
        return pd.DataFrame(columns=['__KEY__'])

    # 2. 查找合同列 (从字典中获取元组)
    contract_col_kw, contract_exact = required_cols_dict.get('合同', ('合同', False))
    contract_col = find_col(df_concat, contract_col_kw, exact=contract_exact)
    
    if not contract_col:
        st.warning(f"⚠️ 在 {prefix} 参考表中未找到'合同'列 (关键字: '{contract_col_kw}', 精确: {contract_exact})，跳过此数据源。")
        return pd.DataFrame(columns=['__KEY__'])
        
    # 3. 提取列 & 重命名
    cols_to_extract = [contract_col]
    col_mapping = {} # '原始列名' -> 'ref_prefix_标准名'
    
    for std_name, (col_kw, is_exact) in required_cols_dict.items(): # <--- V2: 解包元组
        if std_name == '合同': continue 
            
        actual_col = find_col(df_concat, col_kw, exact=is_exact) # <--- V2: 使用 is_exact
        
        if actual_col:
            cols_to_extract.append(actual_col)
            col_mapping[actual_col] = f"ref_{prefix}_{std_name}"
        else:
            # V2: 提供更详细的警告
            st.warning(f"⚠️ 在 {prefix} 参考表中未找到列 (关键字: '{col_kw}', 精确: {is_exact})")
            
    if len(cols_to_extract) == 1: 
        st.warning(f"⚠️ 在 {prefix} 参考表中未找到任何所需字段，跳过。")
        return pd.DataFrame(columns=['__KEY__'])

    # 4. 创建标准DF
    std_df = df_concat[list(set(cols_to_extract))].copy()
    std_df['__KEY__'] = normalize_contract_key(std_df[contract_col])
    std_df = std_df.rename(columns=col_mapping)
    
    final_cols = ['__KEY__'] + list(col_mapping.values())
    
    # 确保 final_cols 都在 std_df 中
    final_cols_in_df = [col for col in final_cols if col in std_df.columns]
    std_df = std_df[final_cols_in_df]
    
    # 5. 去重
    std_df = std_df.drop_duplicates(subset=['__KEY__'], keep='first')
    return std_df

def compare_series_vec(s_main, s_ref, compare_type='text', tolerance=0, multiplier=1):
    """
    (新) 向量化比较函数，复刻所有业务逻辑。
    """
    # 0. 识别 Merge 失败
    merge_failed_mask = s_ref.isna()

    # 1. 预处理空值
    main_is_na = pd.isna(s_main) | (s_main.astype(str).str.strip().isin(["", "nan", "None"]))
    ref_is_na = pd.isna(s_ref) | (s_ref.astype(str).str.strip().isin(["", "nan", "None"]))
    both_are_na = main_is_na & ref_is_na
    
    errors = pd.Series(False, index=s_main.index)

    # 2. 日期比较
    if compare_type == 'date':
        d_main = pd.to_datetime(s_main, errors='coerce').dt.normalize()
        d_ref = pd.to_datetime(s_ref, errors='coerce').dt.normalize()
        
        valid_dates_mask = d_main.notna() & d_ref.notna()
        date_diff_mask = (d_main != d_ref)
        errors = valid_dates_mask & date_diff_mask
        
        one_is_date_one_is_not = (d_main.notna() & d_ref.isna() & ~ref_is_na) | \
                                 (d_main.isna() & ~main_is_na & d_ref.notna())
        errors |= one_is_date_one_is_not

    # 3. 数值比较
    elif compare_type == 'num' or compare_type == 'rate' or compare_type == 'term':
        s_main_norm = s_main.apply(normalize_num)
        s_ref_norm = s_ref.apply(normalize_num)
                   
        # 特殊：期限（乘数）
        if compare_type == 'term':
            s_ref_norm = pd.to_numeric(s_ref_norm, errors='coerce') * multiplier

        is_num_main = s_main_norm.apply(lambda x: isinstance(x, (int, float)))
        is_num_ref = s_ref_norm.apply(lambda x: isinstance(x, (int, float)))
        both_are_num = is_num_main & is_num_ref

        if both_are_num.any():
            diff = (s_main_norm[both_are_num] - s_ref_norm[both_are_num]).abs()
            errors.loc[both_are_num] = (diff > (tolerance + 1e-6))
            
        one_is_num_one_is_not = (is_num_main & ~is_num_ref & ~ref_is_na) | \
                                (~is_num_main & ~main_is_na & is_num_ref)
        errors |= one_is_num_one_is_not

    # 4. 文本比较
    else: # compare_type == 'text'
        s_main_norm_text = s_main.apply(normalize_text)
        s_ref_norm_text = s_ref.apply(normalize_text)
        errors = (s_main_norm_text != s_ref_norm_text)

    # 5. 最终错误逻辑
    final_errors = errors & ~both_are_na
    lookup_failure_mask = merge_failed_mask & ~main_is_na
    final_errors = final_errors & ~lookup_failure_mask
    
    return final_errors

# ========== 读取文件 & 预处理 ==========
st.info("ℹ️ 正在读取并预处理所有文件...")

tc_file = find_file(uploaded_files, "提成")
fk_file = find_file(uploaded_files, "放款明细")
ec_file = find_file(uploaded_files, "二次明细")
original_file = find_file(uploaded_files, "原表")

# 1. 读取主表 (提成)
# ... (这部分不变, 保持原样) ...
tc_xls = pd.ExcelFile(tc_file)
sheet_total = next((s for s in tc_xls.sheet_names if "总" in s), None)
sheets_qk = [s for s in tc_xls.sheet_names if "轻卡" in s]
sheets_zk = [s for s in tc_xls.sheet_names if "重卡" in s]

tc_sheets = {
    "总": [pd.read_excel(tc_file, sheet_name=sheet_total)] if sheet_total else [],
    "轻卡": [pd.read_excel(tc_file, sheet_name=s) for s in sheets_qk],
    "重卡": [pd.read_excel(tc_file, sheet_name=s) for s in sheets_zk],
}

# 2. 读取并预处理参考表
# --- 放款明细 (fk) ---
fk_xls = pd.ExcelFile(fk_file)
fk_dfs_raw = [pd.read_excel(fk_file, sheet_name=s) for s in fk_xls.sheet_names if "潮掣" in s]

# --- VVVV (【核心修改】更新为 (关键字, 是否精确) 的元组) VVVV ---
fk_cols_needed = {
    # {标准名: (关键字, 精确匹配)}
    '合同': ('合同', False),
    '放款日期': ('放款日期', False),
    '提报人员': ('提报人员', False),
    '城市经理': ('城市经理', False),
    '租赁本金': ('租赁本金', False),
    'xirr': ('xirr', False),
    '租赁期限/年': ('租赁期限/年', False), # 模糊匹配也OK，但精确更好
    '家访': ('家访', False),
    '类型': ('类型', True),
    '放款金额': ('放款金额', False)# <--- 这就是您要的修复！
}
fk_std = prepare_ref_df(fk_dfs_raw, fk_cols_needed, "fk")

# --- 二次明细 (ec) ---
ec_xls = pd.ExcelFile(ec_file)
ec_dfs_raw = [pd.read_excel(ec_file, sheet_name=s) for s in ec_xls.sheet_names]
ec_cols_needed = {
    '合同': ('合同', False),
    '出本流程时间': ('出本流程时间', False)
}
ec_std = prepare_ref_df(ec_dfs_raw, ec_cols_needed, "ec")

# --- 原表 (original) ---
original_dfs_raw = [pd.read_excel(original_file)]
original_cols_needed = {
    '合同': ('合同', False),
    '年化nim': ('年化nim', False)
}
orig_std = prepare_ref_df(original_dfs_raw, original_cols_needed, "original")
# --- ^^^^ (修改结束) ^^^^ ---

all_std_dfs = {
    "fk": fk_std,
    "ec": ec_std,
    "orig": orig_std
}

st.success(f"✅ 提成表已读取：总({len(tc_sheets['总'])})、轻卡({len(tc_sheets['轻卡'])})、重卡({len(tc_sheets['重卡'])})")
st.success("✅ 所有参考文件已预处理完成。")

# 3. 定义 MAPPING (保持不变)
MAPPING = {
    "放款日期": ("放款明细", "放款日期", 0, 1),
    "提报人员": ("放款明细", "提报人员", 0, 1),
    "城市经理": ("放款明细", "城市经理", 0, 1),
    "租赁本金": ("放款明细", "租赁本金", 0, 1),
    "收益率": ("放款明细", "xirr", 0.005, 1), 
    "期限": ("放款明细", "租赁期限/年", 0.5, 12),
    "家访": ("放款明细", "家访", 0, 1),
    "人员类型": ("放款明细", "类型", 0, 1),
    "二次交接": ("二次明细", "出本流程时间", 0, 1),
    "计算提成金额": ("放款明细", "放款金额", 0, 1)
}

# =====================================
# 🧮 核心审核函数 (向量化版)
# =====================================
def audit_one_sheet_vec(tc_df, sheet_label, all_std_dfs):
    contract_col_main = find_col(tc_df, "合同")
    if not contract_col_main:
        st.warning(f"⚠️ {sheet_label}：未找到‘合同’列，跳过。")
        return None, None, 0, 0
    
    # 1. 准备主表
    tc_df['__ROW_IDX__'] = tc_df.index
    tc_df['__KEY__'] = normalize_contract_key(tc_df[contract_col_main])

    # 2. 一次性合并所有参考数据
    merged_df = tc_df.copy()
    for std_df in all_std_dfs.values():
        if not std_df.empty:
            merged_df = pd.merge(merged_df, std_df, on='__KEY__', how='left')

    # 3. === 遍历字段进行向量化比对 ===
    total_errors = 0
    errors_locations = set() # 存储 (row_idx, col_name)
    row_has_error = pd.Series(False, index=merged_df.index)

    progress = st.progress(0)
    status = st.empty()
    n = len(merged_df)

    for i, (main_kw, (src, ref_kw, tol, mult)) in enumerate(MAPPING.items()):
        
        exact_main = "期限" in main_kw or main_kw == "人员类型"
        main_col = find_col(merged_df, main_kw, exact=exact_main)
        if not main_col:
            continue
            
        status.text(f"{sheet_label} 审核进度：{i+1}/{len(MAPPING)} - {main_kw}")
        
        s_main = merged_df[main_col]
        
        # 4. === (核心) 处理条件逻辑 ===
        if main_kw == "收益率":
            person_type_col = find_col(merged_df, "人员类型", exact=True)
            if not person_type_col:
                continue # 无法判断类型，跳过
                
            s_ref_fk = merged_df.get('ref_fk_xirr') # 放款明细
            s_ref_orig = merged_df.get('ref_orig_年化nim') # 原表
            
            # (健壮性检查: 如果 'xirr' 列不存在，则创建一个空的 Series)
            if s_ref_fk is None:
                s_ref_fk = pd.Series(pd.NA, index=merged_df.index)
            
            # 默认使用放款明细
            s_ref_final = s_ref_fk.copy()
            
            # 如果类型为"轻卡", 则覆盖为"原表"的值
            if s_ref_orig is not None:
                # --- VVVV (【核心修复】使用 normalize_text) VVVV ---
                person_type_normalized = merged_df[person_type_col].apply(normalize_text)
                mask_light_truck = (person_type_normalized == "轻卡") # '轻卡' 已经是小写
                # --- ^^^^ (修复结束) ^^^^ ---
                s_ref_final.loc[mask_light_truck] = s_ref_orig.loc[mask_light_truck]
            
            errors_mask = compare_series_vec(s_main, s_ref_final, compare_type='rate', tolerance=tol)
        
        elif "日期" in main_kw or main_kw == "二次交接":
            ref_col_name = f"ref_{'ec' if src == '二次明细' else 'fk'}_{ref_kw}"
            s_ref = merged_df.get(ref_col_name)
            errors_mask = compare_series_vec(s_main, s_ref, compare_type='date')
            
        elif "期限" in main_kw:
            ref_col_name = f"ref_fk_{ref_kw}"
            s_ref = merged_df.get(ref_col_name)
            errors_mask = compare_series_vec(s_main, s_ref, compare_type='term', tolerance=tol, multiplier=mult)

        elif main_kw in ["租赁本金", "家访"]: # 其他数值
            ref_col_name = f"ref_fk_{ref_kw}"
            s_ref = merged_df.get(ref_col_name)
            errors_mask = compare_series_vec(s_main, s_ref, compare_type='num', tolerance=tol)

        else: # 文本
            ref_col_name = f"ref_fk_{ref_kw}"
            s_ref = merged_df.get(ref_col_name)
            errors_mask = compare_series_vec(s_main, s_ref, compare_type='text')
            
        # 5. 累积错误
        if errors_mask is not None and errors_mask.any():
            total_errors += errors_mask.sum()
            row_has_error |= errors_mask
            
            bad_indices = merged_df[errors_mask]['__ROW_IDX__']
            for idx in bad_indices:
                errors_locations.add((idx, main_col))
                
        progress.progress((i + 1) / len(MAPPING))

    status.text(f"{sheet_label} 比对完成，正在生成标注文件...")

    # 6. === 快速写入 Excel (替换旧的循环) ===
    wb = Workbook()
    ws = wb.active
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # 准备原始列
    original_cols_list = list(tc_df.drop(columns=['__ROW_IDX__', '__KEY__']).columns)
    col_name_to_idx = {name: i + 1 for i, name in enumerate(original_cols_list)}

    # 写入表头 + 数据
    for r in dataframe_to_rows(merged_df[original_cols_list], index=False, header=True):
        ws.append(r)
    
    # 标红
    for (row_idx, col_name) in errors_locations:
        if col_name in col_name_to_idx:
            excel_row = row_idx + 2 # (row_idx 0-based) + (1 for header) + (1 for 1-based)
            excel_col = col_name_to_idx[col_name]
            ws.cell(excel_row, excel_col).fill = red_fill
            
    # 标黄
    if contract_col_main in col_name_to_idx:
        contract_col_excel_idx = col_name_to_idx[contract_col_main]
        error_row_indices = merged_df[row_has_error]['__ROW_IDX__']
        for row_idx in error_row_indices:
            excel_row = row_idx + 2
            ws.cell(excel_row, contract_col_excel_idx).fill = yellow_fill

    output_full = BytesIO()
    wb.save(output_full)
    output_full.seek(0)
    
    # 7. === (新) 快速生成精简错误表 ===
    output_err = BytesIO()
    error_row_count = row_has_error.sum()
    
    if error_row_count > 0:
        try:
            df_errors_only = merged_df.loc[row_has_error, original_cols_list].copy()
            
            original_indices_with_error = merged_df.loc[row_has_error, '__ROW_IDX__']
            original_idx_to_new_excel_row = {
                original_idx: new_row_num 
                for new_row_num, original_idx in enumerate(original_indices_with_error, start=2)
            }

            wb_err = Workbook()
            ws_err = wb_err.active
            
            for r in dataframe_to_rows(df_errors_only, index=False, header=True):
                ws_err.append(r)
                
            for (original_row_idx, col_name) in errors_locations:
                if original_row_idx in original_idx_to_new_excel_row:
                    new_row = original_idx_to_new_excel_row[original_row_idx]
                    if col_name in col_name_to_idx:
                        new_col = col_name_to_idx[col_name]
                        ws_err.cell(row=new_row, column=new_col).fill = red_fill
            
            # 标黄合同号
            contract_col_excel_idx = col_name_to_idx[contract_col_main]
            for new_row_num in original_idx_to_new_excel_row.values():
                 ws_err.cell(row=new_row_num, column=contract_col_excel_idx).fill = yellow_fill

            wb_err.save(output_err)
            output_err.seek(0)
        except Exception as e:
            st.error(f"❌ 生成“错误精简版”文件时出错: {e}")
            output_err = None # 设为None
    else:
        output_err = None # 没有错误
        
    return output_full, output_err, total_errors, error_row_count
    
# ========== 审核所有 sheet ==========
results = {}
for label, df_list in tc_sheets.items():
    if not df_list:
        continue
    for i, df in enumerate(df_list, start=1):
        tag = f"{label}{i if len(df_list) > 1 else ''}"
        st.divider()
        st.subheader(f"📘 正在审核：{tag}")
        full, err, errs, rows = audit_one_sheet_vec(df, tag, all_std_dfs)
        results[tag] = (full, err, errs, rows)

# ========== 🔍 反向漏填检查（使用标准化Key） ==========
st.divider()
st.subheader("🔍 反向漏填检查（仅基于放款明细中包含“潮掣”的sheet）")

# 1. 从 "总" sheet 获取标准合同号
contracts_total = set()
if tc_sheets["总"]:
    df_total = tc_sheets["总"][0]
    col = find_col(df_total, "合同", exact=False)
    if col is not None:
        contracts_total = set(normalize_contract_key(df_total[col].dropna()))

# 2. 从预处理的 fk_std DataFrame 获取标准合同号
# (这比重新读取 fk_dfs 要快得多，且已标准化)
contracts_fk = set(fk_std['__KEY__'].dropna())

missing_contracts = sorted(list(contracts_fk - contracts_total))

if missing_contracts:
    st.warning(f"⚠️ 发现 {len(missing_contracts)} 个合同号存在于放款明细中，但未出现在提成表‘总’sheet中")
    df_missing = pd.DataFrame({"漏填合同号": missing_contracts})
    output_missing = BytesIO()
    
    # (使用 openpyxl 写入，避免额外的 pd.ExcelWriter 依赖)
    wb_miss = Workbook()
    ws_miss = wb_miss.active
    ws_miss.cell(1, 1, "漏填合同号")
    for r, contract in enumerate(missing_contracts, start=2):
        ws_miss.cell(r, 1, contract)
        
    wb_miss.save(output_missing)
    output_missing.seek(0)
    
    st.download_button(
        "📥 下载漏填合同号表（基于放款明细-潮掣）",
        data=output_missing,
        file_name="提成_漏填合同号_基于放款明细_潮掣.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
else:
    st.success("✅ 未发现漏填合同号（基于放款明细-潮掣）。")

# ========== 下载区 ==========
st.divider()
st.subheader("📤 下载审核结果文件")

for tag, (full, err, errs, rows) in results.items():
    st.write(f"📘 **{tag}**：发现 {errs} 个错误，共 {rows} 行异常")
    st.download_button(
        f"📥 下载 {tag} 审核标注版",
        data=full,
        file_name=f"提成_{tag}_审核标注版.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key=f"download_full_{tag}" # 确保Key唯一
    )
    
    # --- VVVV (添加检查) VVVV ---
    if rows > 0 and err is not None:
    # --- ^^^^ (添加检查) ^^^^ ---
        st.download_button(
            f"📥 下载 {tag} 错误精简版（含红黄标记）",
            data=err,
            file_name=f"提成_{tag}_错误精简版.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"download_err_{tag}" # 确保Key唯一
        )

st.success("✅ 所有sheet审核完成！")
